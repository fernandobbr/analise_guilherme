"""
Converte Dados1.xlsx → dados.js
Execute: python gerar_dados.py
"""
import openpyxl
import json

XLSX = r'Dados1.xlsx'
OUT  = r'dados.js'

DAY_NAMES = {0:'Segunda', 1:'Terca', 2:'Quarta', 3:'Quinta', 4:'Sexta'}

def clean_nome(s):
    if not s: return None
    s = str(s).strip()
    idx = s.find(' - ')
    return s[:idx].strip() if idx >= 0 else s

def time_to_h(t):
    if t is None: return 0.0
    if hasattr(t, 'hour'):
        return round(t.hour + t.minute/60 + t.second/3600, 4)
    return 0.0

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── 1. Colaboradores ──────────────────────────────────────────────
ws_c = wb['Dados_Colaboradores']
colabs = []
mat_lookup = {}   # mat → {setor, turno}

for r in ws_c.iter_rows(min_row=2, values_only=True):
    if not r[1] and not r[0]: continue
    if not r[2]: continue
    mat    = int(r[0]) if r[0] else None
    nome   = clean_nome(r[1])
    setor  = str(r[2]).strip() if r[2] else None
    turno  = str(r[3]).strip() if r[3] else None
    status = str(r[4]).strip() if r[4] else None
    colabs.append({'mat': mat, 'nome': nome, 'setor': setor, 'turno': turno, 'status': status})
    if mat:
        mat_lookup[mat] = {'setor': setor, 'turno': turno}

# ── 2. Registros diários (Seg–Sex) ────────────────────────────────
ws_d = wb['dados']
registros = []

for r in ws_d.iter_rows(min_row=2, values_only=True):
    data, mat = r[0], r[1]
    if not data or not hasattr(data, 'weekday') or not mat: continue
    wd = data.weekday()
    if wd not in DAY_NAMES: continue
    o = str(r[7]).strip() if r[7] and str(r[7]).strip() not in ('', '-', 'None') else ''
    registros.append({
        'm': int(mat),
        'd': DAY_NAMES[wd],
        'p': time_to_h(r[2]),
        't': time_to_h(r[3]),
        'x': round(float(r[8]) if r[8] else 0.0, 4),
        'o': o,
    })

# ── 3. HE Sábado (semana mais recente) ───────────────────────────
ws_he = wb['HE_SABADO']
rows_he = [r for r in ws_he.iter_rows(min_row=2, values_only=True) if r[0]]
sem_max = max(r[0] for r in rows_he) if rows_he else None
he_sabado = []

for r in rows_he:
    if r[0] != sem_max: continue
    mat   = int(r[1]) if r[1] else None
    nome  = str(r[2]).strip() if r[2] else ''
    s     = str(r[4]).upper().strip() == 'SIM' if r[4] else False
    h     = int(r[6].total_seconds()) if r[6] and hasattr(r[6], 'total_seconds') else 0
    info  = mat_lookup.get(mat, {})
    he_sabado.append({
        'mat':   mat,
        'n':     nome,
        'setor': info.get('setor'),
        'turno': info.get('turno'),
        's':     s,
        'h':     h,
    })

# ── Gera dados.js ─────────────────────────────────────────────────
dados = {'colabs': colabs, 'registros': registros, 'heSabado': he_sabado}
js = ('/* AUTO-GERADO de Dados1.xlsx — nao editar manualmente */\n'
      'const DADOS_XLSX = ' + json.dumps(dados, ensure_ascii=False) + ';')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(js)

print(f'OK: {len(colabs)} colabs | {len(registros)} registros | {len(he_sabado)} HE (semana {sem_max})')
